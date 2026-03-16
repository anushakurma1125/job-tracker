import os
import io
import json
import threading
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from werkzeug.security import generate_password_hash, check_password_hash

from database import (init_db, add_job, get_jobs, get_job, update_job, delete_job,
                       job_exists, get_existing_links, bulk_add_jobs,
                       add_resume, get_resumes, get_resume_file, update_resume_label, delete_resume,
                       get_email_settings, save_email_settings, get_gmail_credentials,
                       update_last_scanned, delete_email_settings,
                       get_earliest_applied_date, get_active_jobs,
                       add_scan_log, get_scan_logs,
                       save_scan_checkpoint, clear_scan_checkpoint,
                       create_user, get_user_by_username,
                       get_admin_stats,
                       create_access_request, get_access_requests,
                       approve_access_request, reject_access_request,
                       get_request_by_token, mark_token_used)
from extractor import extract_job_details

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")

# Always initialize the database (works with both direct run and Gunicorn)
init_db()

# ── Background scan state (per-user, in-memory) ──
_scan_state = {}   # {user_id: {"status": "scanning"|"done"|"error", "progress": str, "result": dict}}
_scan_lock = threading.Lock()


def _background_scan(uid, creds_json, jobs, since_date, is_first_scan):
    """Run email scan in a background thread (no HTTP timeout constraint)."""
    try:
        from email_scanner import scan_for_rejections

        def on_progress(msg):
            with _scan_lock:
                if uid in _scan_state:
                    _scan_state[uid]["progress"] = msg

        result = scan_for_rejections(
            creds_json, jobs, since_date,
            is_first_scan=is_first_scan,
            progress_callback=on_progress,
        )

        # Update matched jobs to Rejected
        for match in result.get("details", []):
            update_job(match["job_id"], {
                "status": "Rejected",
                "comment": f"Auto-detected rejection from email: {match.get('email_subject', '')[:80]}",
            }, uid)

        # Save scan log
        add_scan_log(
            result["emails_checked"],
            result["rejections_found"],
            json.dumps(result["details"]),
            uid,
        )

        # Mark scan complete
        update_last_scanned(uid)
        clear_scan_checkpoint(uid)

        with _scan_lock:
            _scan_state[uid] = {"status": "done", "result": result}

    except Exception as e:
        import traceback
        traceback.print_exc()
        with _scan_lock:
            _scan_state[uid] = {"status": "error", "result": {"error": f"Scan failed: {str(e)}"}}


@app.errorhandler(Exception)
def handle_exception(e):
    """Return JSON for API errors instead of HTML error pages."""
    if request.path.startswith("/api/"):
        return jsonify({"error": str(e)}), 500
    return e


@app.errorhandler(500)
def handle_500(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "Internal server error"}), 500
    return e


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def current_user_id():
    return session["user_id"]


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = get_user_by_username(username)
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["display_name"] = user["display_name"] or user["username"]
            return redirect(url_for("index"))
        return render_template("login.html", error="Invalid username or password")
    return render_template("login.html")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if session.get("user_id"):
        return redirect(url_for("index"))

    token = request.args.get("token", "") or request.form.get("token", "")
    if not token:
        return render_template("signup.html", error="An invite link is required to create an account.", no_form=True)

    invite = get_request_by_token(token)
    if not invite:
        return render_template("signup.html", error="Invalid or expired invite link.", no_form=True)

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        display_name = request.form.get("display_name", "").strip()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")

        if not username or not password:
            return render_template("signup.html", error="Username and password are required", token=token, invite=invite)
        if len(password) < 6:
            return render_template("signup.html", error="Password must be at least 6 characters", token=token, invite=invite)
        if password != confirm:
            return render_template("signup.html", error="Passwords do not match", token=token, invite=invite)

        existing = get_user_by_username(username)
        if existing:
            return render_template("signup.html", error="Username already taken", token=token, invite=invite)

        password_hash = generate_password_hash(password)
        user = create_user(username, password_hash, display_name or invite.get("name", username))
        mark_token_used(token)

        session["user_id"] = user["id"]
        session["display_name"] = user["display_name"]
        return redirect(url_for("index"))
    return render_template("signup.html", token=token, invite=invite)


@app.route("/request-access", methods=["GET", "POST"])
def request_access():
    if session.get("user_id"):
        return redirect(url_for("index"))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        about = request.form.get("about", "").strip()

        if not name or not email:
            return render_template("request_access.html", error="Name and email are required")

        create_access_request(name, email, about)

        # Try sending admin notification email
        _send_access_request_email(name, email, about)

        return render_template("request_access.html", success=True)
    return render_template("request_access.html")


def _send_access_request_email(name, email, about):
    """Send notification email to admin. Fails silently if not configured."""
    admin_email = os.environ.get("ADMIN_EMAIL", "")
    app_password = os.environ.get("GMAIL_APP_PASSWORD", "")
    if not admin_email or not app_password:
        return
    try:
        import smtplib
        from email.mime.text import MIMEText
        msg = MIMEText(
            f"New access request for Job Tracker:\n\n"
            f"Name: {name}\nEmail: {email}\n\nAbout:\n{about}\n\n"
            f"Go to your admin dashboard to approve or reject."
        )
        msg["Subject"] = f"Job Tracker Access Request — {name}"
        msg["From"] = admin_email
        msg["To"] = admin_email
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(admin_email, app_password)
            s.send_message(msg)
    except Exception:
        pass  # Email is optional — request is saved to DB regardless


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    return render_template("index.html", display_name=session.get("display_name", ""))


@app.route("/admin-dashboard")
def serve_admin_dashboard():
    return send_file("admin_dashboard.html")


@app.route("/api/jobs", methods=["GET"])
@login_required
def list_jobs():
    status = request.args.get("status")
    jobs = get_jobs(current_user_id(), status=status)
    return jsonify(jobs)


@app.route("/api/jobs", methods=["POST"])
@login_required
def create_job():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Invalid request body"}), 400

        link = data.get("link", "").strip()
        if not link:
            return jsonify({"error": "Job link is required"}), 400

        if job_exists(link, current_user_id()):
            return jsonify({"error": "This job link has already been added"}), 409

        extracted = extract_job_details(link)
        job = add_job(extracted, current_user_id())
        return jsonify(job), 201
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/jobs/bulk", methods=["POST"])
@login_required
def bulk_upload():
    from openpyxl import load_workbook
    import io

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an Excel file (.xlsx)"}), 400

    try:
        from datetime import datetime as dt
        wb = load_workbook(io.BytesIO(file.read()), read_only=True)
        ws = wb.active

        # Read header row to map columns
        headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        col_map = {}
        for i, h in enumerate(headers):
            if "applied" in h and "date" in h:
                col_map["applied_date"] = i
            elif "company" in h:
                col_map["company"] = i
            elif "role" in h or "title" in h or "position" in h:
                col_map["role"] = i
            elif "posted" in h:
                col_map["posted_on"] = i
            elif "description" in h:
                col_map["job_description"] = i
            elif "link" in h or "url" in h:
                col_map["link"] = i
            elif "status" in h:
                col_map["status"] = i
            elif "comment" in h or "note" in h:
                col_map["comment"] = i
            elif "visa" in h:
                col_map["visa_answer"] = i

        # First pass: parse all rows from the spreadsheet
        parsed_rows = []
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                def get_val(key):
                    if key in col_map and col_map[key] < len(row):
                        val = row[col_map[key]]
                        if val is None:
                            return ""
                        if hasattr(val, 'strftime'):
                            return val.strftime("%Y-%m-%d")
                        return str(val).strip()
                    return ""

                job_data = {
                    "applied_date": get_val("applied_date"),
                    "company": get_val("company"),
                    "role": get_val("role"),
                    "posted_on": get_val("posted_on"),
                    "job_description": get_val("job_description"),
                    "link": get_val("link"),
                    "status": get_val("status") or "Applied",
                    "comment": get_val("comment"),
                    "visa_answer": get_val("visa_answer"),
                }

                if not job_data["company"] and not job_data["role"] and not job_data["link"]:
                    continue

                parsed_rows.append(job_data)
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        wb.close()

        # Single DB call: check all existing links at once
        uid = current_user_id()
        all_links = [r["link"] for r in parsed_rows if r["link"]]
        existing_links = get_existing_links(all_links, uid) if all_links else set()

        # Split into new vs duplicate
        to_insert = []
        skipped = 0
        for row_data in parsed_rows:
            if row_data["link"] and row_data["link"] in existing_links:
                skipped += 1
            else:
                to_insert.append(row_data)

        # Single DB call: batch insert all new jobs
        added = bulk_add_jobs(to_insert, uid)

        return jsonify({
            "added": added,
            "skipped": skipped,
            "errors": errors,
        }), 200

    except Exception as e:
        return jsonify({"error": f"Failed to process file: {str(e)}"}), 500


@app.route("/api/jobs/<int:job_id>", methods=["GET"])
@login_required
def get_single_job(job_id):
    job = get_job(job_id, current_user_id())
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job)


@app.route("/api/jobs/<int:job_id>", methods=["PUT"])
@login_required
def update_single_job(job_id):
    data = request.get_json()
    job = update_job(job_id, data, current_user_id())
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job)


@app.route("/api/jobs/<int:job_id>", methods=["DELETE"])
@login_required
def delete_single_job(job_id):
    delete_job(job_id, current_user_id())
    return jsonify({"message": "Job deleted"}), 200


# ── Resume Routes ──

@app.route("/api/resumes", methods=["GET"])
@login_required
def list_resumes():
    resumes = get_resumes(current_user_id())
    return jsonify(resumes)


@app.route("/api/resumes", methods=["POST"])
@login_required
def upload_resume():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files are allowed"}), 400
    label = request.form.get("label", "").strip()
    file_data = file.read()
    file_size = len(file_data)
    if file_size > 5 * 1024 * 1024:
        return jsonify({"error": "File too large (max 5 MB)"}), 400
    if file_size == 0:
        return jsonify({"error": "File is empty"}), 400
    try:
        resume = add_resume(file.filename, label, file_data, file_size, current_user_id())
        return jsonify(resume), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/resumes/<int:resume_id>/download", methods=["GET"])
@login_required
def download_resume(resume_id):
    row = get_resume_file(resume_id, current_user_id())
    if not row:
        return jsonify({"error": "Resume not found"}), 404
    file_data = row["file_data"]
    if isinstance(file_data, memoryview):
        file_data = bytes(file_data)
    return send_file(
        io.BytesIO(file_data),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=row["filename"],
    )


@app.route("/api/resumes/<int:resume_id>", methods=["PUT"])
@login_required
def update_single_resume(resume_id):
    data = request.get_json()
    label = data.get("label", "").strip()
    resume = update_resume_label(resume_id, label, current_user_id())
    if not resume:
        return jsonify({"error": "Resume not found"}), 404
    return jsonify(resume)


@app.route("/api/resumes/<int:resume_id>", methods=["DELETE"])
@login_required
def delete_single_resume(resume_id):
    delete_resume(resume_id, current_user_id())
    return jsonify({"message": "Resume deleted"}), 200


# ── Email Settings & Scan Routes ──

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]


@app.route("/api/settings/email", methods=["GET"])
@login_required
def get_email_status():
    settings = get_email_settings(current_user_id())
    if not settings or not settings.get("enabled"):
        return jsonify({"connected": False, "configured": bool(GOOGLE_CLIENT_ID)})
    return jsonify({
        "connected": True,
        "configured": True,
        "email": settings.get("gmail_email", ""),
        "last_scanned_at": settings.get("last_scanned_at", ""),
    })


@app.route("/api/settings/email/auth-url", methods=["GET"])
@login_required
def get_auth_url():
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        return jsonify({"error": "Google OAuth credentials not configured. Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET environment variables."}), 400

    from google_auth_oauthlib.flow import Flow

    # Determine the redirect URI based on the request
    redirect_uri = request.url_root.rstrip("/") + "/api/settings/email/oauth-callback"

    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [redirect_uri],
            }
        },
        scopes=GMAIL_SCOPES,
        redirect_uri=redirect_uri,
    )

    auth_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
    )
    session["oauth_state"] = state
    return jsonify({"auth_url": auth_url})


@app.route("/api/settings/email/oauth-callback")
@login_required
def oauth_callback():
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        return "OAuth not configured", 400

    from google_auth_oauthlib.flow import Flow

    redirect_uri = request.url_root.rstrip("/") + "/api/settings/email/oauth-callback"

    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [redirect_uri],
            }
        },
        scopes=GMAIL_SCOPES,
        redirect_uri=redirect_uri,
        state=session.get("oauth_state"),
    )

    flow.fetch_token(authorization_response=request.url)
    creds = flow.credentials

    # Get user's email address
    from googleapiclient.discovery import build
    service = build("gmail", "v1", credentials=creds)
    profile = service.users().getProfile(userId="me").execute()
    gmail_email = profile.get("emailAddress", "")

    # Store credentials
    creds_data = {
        "token": creds.token,
        "refresh_token": creds.refresh_token,
        "token_uri": creds.token_uri,
        "client_id": creds.client_id,
        "client_secret": creds.client_secret,
        "scopes": list(creds.scopes) if creds.scopes else GMAIL_SCOPES,
    }
    save_email_settings(json.dumps(creds_data), gmail_email, current_user_id())

    # Redirect back to the app (Settings page)
    return redirect("/#settings-connected")


@app.route("/api/settings/email/disconnect", methods=["POST"])
@login_required
def disconnect_email():
    delete_email_settings(current_user_id())
    return jsonify({"message": "Email disconnected"})


@app.route("/api/settings/email/scan", methods=["POST"])
@login_required
def trigger_scan():
    uid = current_user_id()

    # If a scan is already running, just return the current state
    with _scan_lock:
        if uid in _scan_state and _scan_state[uid].get("status") == "scanning":
            return jsonify({
                "status": "scanning",
                "progress": _scan_state[uid].get("progress", "Scanning..."),
            }), 200

    creds_json = get_gmail_credentials(uid)
    if not creds_json:
        return jsonify({"error": "Gmail not connected. Go to Settings to connect your email."}), 400

    settings = get_email_settings(uid)
    jobs = get_active_jobs(uid)
    if not jobs:
        return jsonify({
            "status": "done",
            "result": {"emails_checked": 0, "rejections_found": 0, "details": [], "message": "No active jobs to check."},
        }), 200

    # Determine scan start date and whether this is the first scan
    is_first_scan = not (settings and settings.get("last_scanned_at"))
    if not is_first_scan:
        since_date = settings["last_scanned_at"]
    else:
        since_date = get_earliest_applied_date(uid)

    if not since_date:
        since_date = "2024-01-01"

    # Start scan in background thread — returns immediately
    with _scan_lock:
        _scan_state[uid] = {"status": "scanning", "progress": "Starting scan..."}

    thread = threading.Thread(
        target=_background_scan,
        args=(uid, creds_json, jobs, since_date, is_first_scan),
        daemon=True,
    )
    thread.start()

    return jsonify({"status": "scanning", "progress": "Starting scan..."}), 200


@app.route("/api/settings/email/scan/status", methods=["GET"])
@login_required
def scan_status():
    uid = current_user_id()
    with _scan_lock:
        state = _scan_state.get(uid)

    if not state:
        return jsonify({"status": "idle"}), 200

    return jsonify(state), 200


# ── ATS Analyzer Routes ──

@app.route("/api/ats/analyze", methods=["POST"])
@login_required
def ats_analyze():
    import anthropic

    data = request.get_json()
    jd = (data.get("jd") or "").strip()
    resume_text = (data.get("resume") or "").strip()
    role = (data.get("role") or "Not specified").strip()
    history = data.get("history") or []

    if not jd or not resume_text:
        return jsonify({"error": "Both job description and resume are required."}), 400

    # Build history context
    ctx = ""
    if history:
        recent = history[-3:]
        ctx = f"Prior analyses ({len(history)}): " + " | ".join(
            f"{h.get('role','?')} score:{h.get('score','?')} gaps:{','.join(h.get('gaps',[]))}"
            for h in recent
        )

    system_prompt = f"""Elite ATS analyzer. {ctx}
Return ONLY compact JSON, no markdown:
{{"verdict":"PASS|LIKELY PASS|BORDERLINE|LIKELY REJECT|REJECT","ats_score":0-100,"apply":{{"decision":"STRONG YES|YES|WORTH A SHOT|RISKY|DON'T BOTHER","confidence":0-100,"reason":"2 sentences max"}},"scores":{{"keywords":0-100,"experience":0-100,"skills":0-100,"formatting":0-100,"impact":0-100}},"matched":["kw"],"missing_critical":["kw"],"missing_nice":["kw"],"padding":["specific padded claim"],"flags":["recruiter red flag"],"genuine":["real strength"],"strengths":["specific strength"],"fixes":["specific fix"],"profile":{{"level":"e.g. Mid-level","gap":"single biggest gap","trajectory":"1 sentence","fit_roles":["role"]}},"pattern":"1 sentence about history pattern or First analysis","summary":"2 sentences honest direct assessment"}}
Rules: be brutally honest, specific to their actual resume, flag keyword stuffing explicitly, no hedging on apply decision."""

    user_msg = f"JOB TITLE: {role}\n\nJOB DESCRIPTION:\n{jd}\n\n---\n\nRESUME:\n{resume_text}"

    try:
        client = anthropic.Anthropic()
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        text = resp.content[0].text if resp.content else ""
        # Strip markdown fences and extract JSON
        import re
        stripped = re.sub(r"```json|```", "", text).strip()
        match = re.search(r"\{[\s\S]*\}", stripped)
        if not match:
            return jsonify({"error": "Failed to parse AI response."}), 500
        result = json.loads(match.group(0))
        return jsonify(result), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Analysis failed: {str(e)}"}), 500


@app.route("/api/ats/rewrites", methods=["POST"])
@login_required
def ats_rewrites():
    import anthropic

    data = request.get_json()
    jd = (data.get("jd") or "").strip()
    resume_text = (data.get("resume") or "").strip()
    role = (data.get("role") or "Not specified").strip()

    if not jd or not resume_text:
        return jsonify({"error": "Both job description and resume are required."}), 400

    system_prompt = """You are a resume rewriter. Given a resume and job description, return ONLY JSON:
{"rewrites":[{"section":"which bullet/section","before":"exact text from resume","after":"rewritten version with metrics and stronger language","why":"one sentence"}]}
Max 4 rewrites. Focus on highest-impact changes. No markdown, no backticks."""

    user_msg = f"JOB TITLE: {role}\n\nJOB DESCRIPTION:\n{jd}\n\n---\n\nRESUME:\n{resume_text}"

    try:
        client = anthropic.Anthropic()
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1500,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        text = resp.content[0].text if resp.content else ""
        import re
        stripped = re.sub(r"```json|```", "", text).strip()
        match = re.search(r"\{[\s\S]*\}", stripped)
        if not match:
            return jsonify({"rewrites": []}), 200
        result = json.loads(match.group(0))
        return jsonify(result), 200
    except Exception:
        return jsonify({"rewrites": []}), 200


@app.after_request
def _cors_admin(response):
    """Add CORS headers for the admin stats endpoint."""
    if request.path == "/api/admin/stats":
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Admin-Key"
        response.headers["Access-Control-Max-Age"] = "3600"
    return response


@app.route("/api/admin/stats", methods=["GET", "OPTIONS"])
def admin_stats():
    if request.method == "OPTIONS":
        return "", 204
    # Auth: either session (user_id=1) or X-Admin-Key header
    admin_key = os.environ.get("ADMIN_DASHBOARD_KEY", "")
    req_key = request.headers.get("X-Admin-Key", "")
    if req_key and admin_key and req_key == admin_key:
        pass  # key-based auth OK
    elif session.get("user_id") == 1:
        pass  # session auth OK
    else:
        return jsonify({"error": "Forbidden"}), 403
    stats = get_admin_stats()
    return jsonify(stats)


@app.route("/api/admin/access-requests", methods=["GET"])
@login_required
def list_access_requests():
    if current_user_id() != 1:
        return jsonify({"error": "Forbidden"}), 403
    return jsonify(get_access_requests())


@app.route("/api/admin/access-requests/<int:req_id>/approve", methods=["POST"])
@login_required
def approve_request(req_id):
    if current_user_id() != 1:
        return jsonify({"error": "Forbidden"}), 403
    token = approve_access_request(req_id)
    base_url = request.url_root.rstrip("/")
    invite_url = f"{base_url}/signup?token={token}"
    return jsonify({"token": token, "invite_url": invite_url})


@app.route("/api/admin/access-requests/<int:req_id>/reject", methods=["POST"])
@login_required
def reject_request(req_id):
    if current_user_id() != 1:
        return jsonify({"error": "Forbidden"}), 403
    reject_access_request(req_id)
    return jsonify({"message": "Request rejected"})


@app.route("/api/settings/email/logs", methods=["GET"])
@login_required
def list_scan_logs():
    logs = get_scan_logs(current_user_id())
    return jsonify(logs)


if __name__ == "__main__":
    app.run(debug=True, port=5001)
